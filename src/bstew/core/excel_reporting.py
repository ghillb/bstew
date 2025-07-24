"""
Excel Reporting Integration for NetLogo BEE-STEWARD v2 Parity
============================================================

Advanced Excel reporting system for comprehensive simulation data export,
analysis, and visualization matching NetLogo's data output capabilities.
"""

from typing import Dict, List, Optional, Any
from pydantic import BaseModel, Field
from enum import Enum
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from datetime import datetime
import os

from .data_collection import ComprehensiveDataCollector
from .colony_mortality import DeadColonyTracker
from .parameter_loader import ParameterLoader


class ReportType(Enum):
    """Types of Excel reports"""

    SIMULATION_SUMMARY = "simulation_summary"
    COLONY_PERFORMANCE = "colony_performance"
    FORAGING_ANALYSIS = "foraging_analysis"
    MORTALITY_REPORT = "mortality_report"
    PARAMETER_ANALYSIS = "parameter_analysis"
    COMPREHENSIVE = "comprehensive"


class ExcelReportGenerator(BaseModel):
    """Comprehensive Excel report generator"""

    model_config = {"validate_assignment": True}

    # Output configuration
    output_directory: str = Field(
        default="artifacts/reports", description="Output directory for reports"
    )
    include_charts: bool = Field(default=True, description="Include charts in reports")
    include_formatting: bool = Field(
        default=True, description="Include advanced formatting"
    )

    # Data sources
    data_collector: Optional[ComprehensiveDataCollector] = Field(
        default=None, description="Data collector"
    )
    mortality_tracker: Optional[DeadColonyTracker] = Field(
        default=None, description="Mortality tracker"
    )
    parameter_loader: Optional[ParameterLoader] = Field(
        default=None, description="Parameter loader"
    )

    # Report configuration
    max_rows_per_sheet: int = Field(
        default=1000000, ge=1000, description="Maximum rows per sheet"
    )
    decimal_places: int = Field(
        default=3, ge=0, le=10, description="Decimal places for numbers"
    )

    def __init__(self, **data: Any) -> None:
        super().__init__(**data)

        # Create output directory if it doesn't exist
        os.makedirs(self.output_directory, exist_ok=True)

        # Excel formatting styles
        self.header_style = {
            "font": Font(bold=True, color="FFFFFF"),
            "fill": PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            ),
            "alignment": Alignment(horizontal="center", vertical="center"),
        }

        self.data_style = {
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

        self.number_style = {"number_format": f"0.{'0' * self.decimal_places}"}

    def generate_simulation_summary_report(
        self, simulation_data: Dict[str, Any]
    ) -> str:
        """Generate comprehensive simulation summary report"""

        workbook = Workbook()

        # Remove default sheet
        if workbook.active:
            workbook.remove(workbook.active)

        # Overview sheet
        self._create_overview_sheet(workbook, simulation_data)

        # Population dynamics sheet
        self._create_population_dynamics_sheet(workbook, simulation_data)

        # Energy analysis sheet
        self._create_energy_analysis_sheet(workbook, simulation_data)

        # Foraging performance sheet
        self._create_foraging_performance_sheet(workbook, simulation_data)

        # Environmental conditions sheet
        self._create_environmental_conditions_sheet(workbook, simulation_data)

        # Summary statistics sheet
        self._create_summary_statistics_sheet(workbook, simulation_data)

        # Save workbook
        filename = f"simulation_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_directory, filename)
        workbook.save(filepath)

        return filepath

    def _create_overview_sheet(
        self, workbook: Workbook, simulation_data: Dict[str, Any]
    ) -> None:
        """Create simulation overview sheet"""

        ws = workbook.create_sheet("Overview")

        # Title
        ws["A1"] = "BSTEW Simulation Overview"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        # Basic information
        row = 3
        info_data = [
            ["Simulation Start Time", simulation_data.get("start_time", "N/A")],
            ["Simulation End Time", simulation_data.get("end_time", "N/A")],
            ["Total Duration (hours)", simulation_data.get("duration_hours", "N/A")],
            ["Total Simulation Steps", simulation_data.get("total_steps", "N/A")],
            ["Colonies Simulated", simulation_data.get("colony_count", "N/A")],
            ["Total Bees Simulated", simulation_data.get("total_bees", "N/A")],
            ["Survival Rate (%)", simulation_data.get("survival_rate", "N/A")],
            ["Data Points Collected", simulation_data.get("data_points", "N/A")],
        ]

        for info in info_data:
            ws[f"A{row}"] = info[0]
            ws[f"B{row}"] = info[1]
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Model parameters section
        row += 2
        ws[f"A{row}"] = "Model Parameters"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        parameters = simulation_data.get("parameters", {})
        for param_name, param_value in parameters.items():
            ws[f"A{row}"] = param_name
            ws[f"B{row}"] = param_value
            row += 1

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=row)

    def _create_population_dynamics_sheet(
        self, workbook: Workbook, simulation_data: Dict[str, Any]
    ) -> None:
        """Create population dynamics analysis sheet"""

        ws = workbook.create_sheet("Population Dynamics")

        # Population time series data
        pop_data = simulation_data.get("population_data", [])

        if pop_data:
            # Create DataFrame
            df = pd.DataFrame(pop_data)

            # Add to worksheet
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create chart if enabled
            if self.include_charts and len(pop_data) > 1:
                self._create_population_chart(ws, len(pop_data))

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=len(pop_data) + 1)

    def _create_energy_analysis_sheet(
        self, workbook: Workbook, simulation_data: Dict[str, Any]
    ) -> None:
        """Create energy analysis sheet"""

        ws = workbook.create_sheet("Energy Analysis")

        # Energy metrics
        energy_data = simulation_data.get("energy_data", [])

        if energy_data:
            # Create DataFrame
            df = pd.DataFrame(energy_data)

            # Add to worksheet
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create energy efficiency chart
            if self.include_charts and len(energy_data) > 1:
                self._create_energy_chart(ws, len(energy_data))

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=len(energy_data) + 1)

    def _create_foraging_performance_sheet(
        self, workbook: Workbook, simulation_data: Dict[str, Any]
    ) -> None:
        """Create foraging performance analysis sheet"""

        ws = workbook.create_sheet("Foraging Performance")

        # Foraging metrics
        foraging_data = simulation_data.get("foraging_data", [])

        if foraging_data:
            # Create DataFrame
            df = pd.DataFrame(foraging_data)

            # Add to worksheet
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create foraging efficiency chart
            if self.include_charts and len(foraging_data) > 1:
                self._create_foraging_chart(ws, len(foraging_data))

        # Summary statistics
        self._add_foraging_summary_statistics(ws, foraging_data)

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=len(foraging_data) + 10)

    def _create_environmental_conditions_sheet(
        self, workbook: Workbook, simulation_data: Dict[str, Any]
    ) -> None:
        """Create environmental conditions sheet"""

        ws = workbook.create_sheet("Environmental Conditions")

        # Environmental data
        env_data = simulation_data.get("environmental_data", [])

        if env_data:
            # Create DataFrame
            df = pd.DataFrame(env_data)

            # Add to worksheet
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create environmental chart
            if self.include_charts and len(env_data) > 1:
                self._create_environmental_chart(ws, len(env_data))

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=len(env_data) + 1)

    def _create_summary_statistics_sheet(
        self, workbook: Workbook, simulation_data: Dict[str, Any]
    ) -> None:
        """Create summary statistics sheet"""

        ws = workbook.create_sheet("Summary Statistics")

        # Calculate and display summary statistics
        stats = simulation_data.get("summary_statistics", {})

        row = 1

        # Population statistics
        ws[f"A{row}"] = "Population Statistics"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 2

        pop_stats = stats.get("population", {})
        for stat_name, stat_value in pop_stats.items():
            ws[f"A{row}"] = stat_name
            ws[f"B{row}"] = stat_value
            row += 1

        row += 1

        # Energy statistics
        ws[f"A{row}"] = "Energy Statistics"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 2

        energy_stats = stats.get("energy", {})
        for stat_name, stat_value in energy_stats.items():
            ws[f"A{row}"] = stat_name
            ws[f"B{row}"] = stat_value
            row += 1

        row += 1

        # Foraging statistics
        ws[f"A{row}"] = "Foraging Statistics"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 2

        foraging_stats = stats.get("foraging", {})
        for stat_name, stat_value in foraging_stats.items():
            ws[f"A{row}"] = stat_name
            ws[f"B{row}"] = stat_value
            row += 1

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=row)

    def _create_population_chart(self, ws: Any, data_rows: int) -> None:
        """Create population dynamics chart"""

        chart = LineChart()
        chart.title = "Population Dynamics Over Time"
        chart.y_axis.title = "Population Count"
        chart.x_axis.title = "Time Step"

        # Assuming columns: Time, Total_Population, Workers, Foragers, etc.
        data = Reference(ws, min_col=2, min_row=1, max_col=5, max_row=data_rows + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, "H2")

    def _create_energy_chart(self, ws: Any, data_rows: int) -> None:
        """Create energy analysis chart"""

        chart = LineChart()
        chart.title = "Energy Levels Over Time"
        chart.y_axis.title = "Energy Units"
        chart.x_axis.title = "Time Step"

        # Assuming columns: Time, Total_Energy, Energy_Intake, Energy_Consumption
        data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=data_rows + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, "H2")

    def _create_foraging_chart(self, ws: Any, data_rows: int) -> None:
        """Create foraging performance chart"""

        chart = LineChart()
        chart.title = "Foraging Performance Over Time"
        chart.y_axis.title = "Efficiency / Success Rate"
        chart.x_axis.title = "Time Step"

        # Assuming columns: Time, Success_Rate, Efficiency, Trips_Per_Day
        data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=data_rows + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, "H2")

    def _create_environmental_chart(self, ws: Any, data_rows: int) -> None:
        """Create environmental conditions chart"""

        chart = LineChart()
        chart.title = "Environmental Conditions Over Time"
        chart.y_axis.title = "Condition Value"
        chart.x_axis.title = "Time Step"

        # Assuming columns: Time, Temperature, Humidity, Resource_Availability
        data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=data_rows + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, "H2")

    def _add_foraging_summary_statistics(
        self, ws: Any, foraging_data: List[Dict[str, Any]]
    ) -> None:
        """Add foraging summary statistics to sheet"""

        if not foraging_data:
            return

        # Calculate summary statistics
        df = pd.DataFrame(foraging_data)

        # Find insertion point
        start_row = len(foraging_data) + 3

        # Add summary title
        ws[f"A{start_row}"] = "Foraging Summary Statistics"
        ws[f"A{start_row}"].font = Font(size=12, bold=True)
        start_row += 2

        # Calculate statistics
        numeric_columns = df.select_dtypes(include=[np.number]).columns

        for col in numeric_columns:
            stats = {
                "Mean": df[col].mean(),
                "Median": df[col].median(),
                "Std Dev": df[col].std(),
                "Min": df[col].min(),
                "Max": df[col].max(),
            }

            ws[f"A{start_row}"] = f"{col} Statistics"
            ws[f"A{start_row}"].font = Font(bold=True)
            start_row += 1

            for stat_name, stat_value in stats.items():
                ws[f"B{start_row}"] = stat_name
                ws[f"C{start_row}"] = round(stat_value, self.decimal_places)
                start_row += 1

            start_row += 1

    def _apply_sheet_formatting(self, ws: Any, max_row: int) -> None:
        """Apply formatting to worksheet"""

        if not self.include_formatting:
            return

        # Header row formatting
        for cell in ws[1]:
            if cell.value:
                cell.font = self.header_style["font"]
                cell.fill = self.header_style["fill"]
                cell.alignment = self.header_style["alignment"]

        # Data formatting
        for row in range(2, max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = self.data_style["alignment"]
                cell.border = self.data_style["border"]

                # Number formatting
                if isinstance(cell.value, (int, float)):
                    cell.number_format = self.number_style["number_format"]

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def generate_colony_performance_report(self, colony_data: Dict[str, Any]) -> str:
        """Generate colony performance report"""

        workbook = Workbook()
        if workbook.active:
            workbook.remove(workbook.active)

        # Individual colony sheets
        for colony_id, data in colony_data.items():
            self._create_colony_sheet(workbook, colony_id, data)

        # Comparative analysis sheet
        self._create_comparative_analysis_sheet(workbook, colony_data)

        # Save workbook
        filename = f"colony_performance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_directory, filename)
        workbook.save(filepath)

        return filepath

    def _create_colony_sheet(
        self, workbook: Workbook, colony_id: str, colony_data: Dict[str, Any]
    ) -> None:
        """Create individual colony performance sheet"""

        ws = workbook.create_sheet(f"Colony_{colony_id}")

        # Colony header
        ws["A1"] = f"Colony {colony_id} Performance Report"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

        # Performance metrics
        row = 3
        metrics = colony_data.get("performance_metrics", {})

        for metric_name, metric_value in metrics.items():
            ws[f"A{row}"] = metric_name
            ws[f"B{row}"] = metric_value
            row += 1

        # Time series data
        time_series = colony_data.get("time_series", [])
        if time_series:
            # Add time series data
            df = pd.DataFrame(time_series)

            start_row = row + 2
            ws[f"A{start_row}"] = "Time Series Data"
            ws[f"A{start_row}"].font = Font(size=12, bold=True)
            start_row += 2

            for r_idx, row_data in enumerate(
                dataframe_to_rows(df, index=False, header=True), start_row
            ):
                for c_idx, value in enumerate(row_data, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create chart
            if self.include_charts and len(time_series) > 1:
                self._create_colony_performance_chart(ws, len(time_series), start_row)

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=row + len(time_series) + 5)

    def _create_colony_performance_chart(
        self, ws: Any, data_rows: int, start_row: int
    ) -> None:
        """Create colony performance chart"""

        chart = LineChart()
        chart.title = "Colony Performance Over Time"
        chart.y_axis.title = "Performance Metric"
        chart.x_axis.title = "Time Step"

        # Assuming first column is time, subsequent columns are metrics
        data = Reference(
            ws, min_col=2, min_row=start_row, max_col=4, max_row=start_row + data_rows
        )
        categories = Reference(
            ws, min_col=1, min_row=start_row + 1, max_row=start_row + data_rows
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, "H10")

    def _create_comparative_analysis_sheet(
        self, workbook: Workbook, colony_data: Dict[str, Any]
    ) -> None:
        """Create comparative analysis sheet"""

        ws = workbook.create_sheet("Comparative Analysis")

        # Title
        ws["A1"] = "Colony Comparative Analysis"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:F1")

        # Create comparison table
        row = 3
        ws[f"A{row}"] = "Colony ID"
        ws[f"B{row}"] = "Survival Rate"
        ws[f"C{row}"] = "Avg Population"
        ws[f"D{row}"] = "Peak Population"
        ws[f"E{row}"] = "Energy Efficiency"
        ws[f"F{row}"] = "Foraging Success"

        # Apply header formatting
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_style["font"]
            cell.fill = self.header_style["fill"]
            cell.alignment = self.header_style["alignment"]

        row += 1

        # Add colony data
        for colony_id, data in colony_data.items():
            metrics = data.get("performance_metrics", {})

            ws[f"A{row}"] = colony_id
            ws[f"B{row}"] = metrics.get("survival_rate", 0)
            ws[f"C{row}"] = metrics.get("average_population", 0)
            ws[f"D{row}"] = metrics.get("peak_population", 0)
            ws[f"E{row}"] = metrics.get("energy_efficiency", 0)
            ws[f"F{row}"] = metrics.get("foraging_success", 0)

            row += 1

        # Create comparison chart
        if self.include_charts and len(colony_data) > 1:
            self._create_comparison_chart(ws, len(colony_data), 4)

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=row)

    def _create_comparison_chart(
        self, ws: Any, num_colonies: int, start_row: int
    ) -> None:
        """Create colony comparison chart"""

        chart = BarChart()
        chart.title = "Colony Performance Comparison"
        chart.y_axis.title = "Performance Value"
        chart.x_axis.title = "Colony ID"

        # Data for comparison
        data = Reference(
            ws,
            min_col=2,
            min_row=start_row - 1,
            max_col=6,
            max_row=start_row + num_colonies - 1,
        )
        categories = Reference(
            ws, min_col=1, min_row=start_row, max_row=start_row + num_colonies - 1
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, "H3")

    def generate_mortality_report(self, mortality_data: Dict[str, Any]) -> str:
        """Generate mortality analysis report"""

        workbook = Workbook()
        if workbook.active:
            workbook.remove(workbook.active)

        # Mortality overview
        self._create_mortality_overview_sheet(workbook, mortality_data)

        # Collapse analysis
        self._create_collapse_analysis_sheet(workbook, mortality_data)

        # Survival factors
        self._create_survival_factors_sheet(workbook, mortality_data)

        # Save workbook
        filename = f"mortality_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_directory, filename)
        workbook.save(filepath)

        return filepath

    def _create_mortality_overview_sheet(
        self, workbook: Workbook, mortality_data: Dict[str, Any]
    ) -> None:
        """Create mortality overview sheet"""

        ws = workbook.create_sheet("Mortality Overview")

        # Title
        ws["A1"] = "Colony Mortality Analysis"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        # Summary statistics
        row = 3
        summary = mortality_data.get("summary_statistics", {})

        summary_data = [
            ["Total Colonies Tracked", summary.get("total_colonies", 0)],
            ["Active Colonies", summary.get("active_colonies", 0)],
            ["Dead Colonies", summary.get("dead_colonies", 0)],
            ["Overall Survival Rate (%)", summary.get("survival_rate", 0) * 100],
            ["Average Lifespan (days)", summary.get("average_lifespan_days", 0)],
            [
                "Average Collapse Duration (days)",
                summary.get("average_collapse_duration_days", 0),
            ],
        ]

        for item in summary_data:
            ws[f"A{row}"] = item[0]
            ws[f"B{row}"] = item[1]
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Collapse reasons
        row += 2
        ws[f"A{row}"] = "Collapse Reasons"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        collapse_reasons = summary.get("collapse_reasons", {})
        for reason, count in collapse_reasons.items():
            ws[f"A{row}"] = reason.replace("_", " ").title()
            ws[f"B{row}"] = count
            row += 1

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=row)

    def _create_collapse_analysis_sheet(
        self, workbook: Workbook, mortality_data: Dict[str, Any]
    ) -> None:
        """Create collapse analysis sheet"""

        ws = workbook.create_sheet("Collapse Analysis")

        # Individual collapse events
        collapse_events = mortality_data.get("collapse_events", [])

        if collapse_events:
            # Create DataFrame
            df = pd.DataFrame(collapse_events)

            # Add to worksheet
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create collapse timeline chart
            if self.include_charts and len(collapse_events) > 1:
                self._create_collapse_timeline_chart(ws, len(collapse_events))

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=len(collapse_events) + 1)

    def _create_collapse_timeline_chart(self, ws: Any, num_events: int) -> None:
        """Create collapse timeline chart"""

        chart = ScatterChart()
        chart.title = "Colony Collapse Timeline"
        chart.y_axis.title = "Days Since Peak Population"
        chart.x_axis.title = "Collapse Event"

        # Assuming column structure includes collapse timeline data
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=num_events + 1)

        chart.add_data(data, titles_from_data=True)

        # Position chart
        ws.add_chart(chart, "H2")

    def _create_survival_factors_sheet(
        self, workbook: Workbook, mortality_data: Dict[str, Any]
    ) -> None:
        """Create survival factors analysis sheet"""

        ws = workbook.create_sheet("Survival Factors")

        # Survival factor analysis
        factors = mortality_data.get("survival_factors", {})

        row = 1
        ws[f"A{row}"] = "Survival Factors Analysis"
        ws[f"A{row}"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        row += 2

        for factor_name, factor_data in factors.items():
            ws[f"A{row}"] = factor_name.replace("_", " ").title()
            ws[f"A{row}"].font = Font(size=12, bold=True)
            row += 1

            if isinstance(factor_data, dict):
                for key, value in factor_data.items():
                    ws[f"B{row}"] = key
                    ws[f"C{row}"] = value
                    row += 1
            else:
                ws[f"B{row}"] = factor_data
                row += 1

            row += 1

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=row)

    def export_comprehensive_report(self, all_data: Dict[str, Any]) -> str:
        """Export comprehensive report with all data"""

        workbook = Workbook()
        if workbook.active:
            workbook.remove(workbook.active)

        # Add all report sections
        self._create_overview_sheet(workbook, all_data.get("simulation_data", {}))
        self._create_population_dynamics_sheet(
            workbook, all_data.get("simulation_data", {})
        )
        self._create_energy_analysis_sheet(
            workbook, all_data.get("simulation_data", {})
        )
        self._create_foraging_performance_sheet(
            workbook, all_data.get("simulation_data", {})
        )
        self._create_environmental_conditions_sheet(
            workbook, all_data.get("simulation_data", {})
        )

        # Colony-specific data
        if "colony_data" in all_data:
            self._create_comparative_analysis_sheet(workbook, all_data["colony_data"])

        # Mortality data
        if "mortality_data" in all_data:
            self._create_mortality_overview_sheet(workbook, all_data["mortality_data"])
            self._create_collapse_analysis_sheet(workbook, all_data["mortality_data"])

        # Parameter analysis
        if "parameter_data" in all_data:
            self._create_parameter_analysis_sheet(workbook, all_data["parameter_data"])

        # Save workbook
        filename = (
            f"comprehensive_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        filepath = os.path.join(self.output_directory, filename)
        workbook.save(filepath)

        return filepath

    def _create_parameter_analysis_sheet(
        self, workbook: Workbook, parameter_data: Dict[str, Any]
    ) -> None:
        """Create parameter analysis sheet"""

        ws = workbook.create_sheet("Parameter Analysis")

        # Title
        ws["A1"] = "Parameter Analysis"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        row = 3

        # Behavioral parameters
        if "behavioral" in parameter_data:
            ws[f"A{row}"] = "Behavioral Parameters"
            ws[f"A{row}"].font = Font(size=14, bold=True)
            row += 1

            behavioral_params = parameter_data["behavioral"]
            for param_name, param_value in behavioral_params.items():
                ws[f"A{row}"] = param_name
                ws[f"B{row}"] = param_value
                row += 1

            row += 1

        # Environmental parameters
        if "environmental" in parameter_data:
            ws[f"A{row}"] = "Environmental Parameters"
            ws[f"A{row}"].font = Font(size=14, bold=True)
            row += 1

            environmental_params = parameter_data["environmental"]
            for param_name, param_value in environmental_params.items():
                ws[f"A{row}"] = param_name
                ws[f"B{row}"] = param_value
                row += 1

            row += 1

        # Colony parameters
        if "colony" in parameter_data:
            ws[f"A{row}"] = "Colony Parameters"
            ws[f"A{row}"].font = Font(size=14, bold=True)
            row += 1

            colony_params = parameter_data["colony"]
            for param_name, param_value in colony_params.items():
                ws[f"A{row}"] = param_name
                ws[f"B{row}"] = param_value
                row += 1

        # Apply formatting
        self._apply_sheet_formatting(ws, max_row=row)

    def get_report_summary(self) -> Dict[str, Any]:
        """Get summary of all generated reports"""

        report_files = []

        # List all Excel files in output directory
        for file in os.listdir(self.output_directory):
            if file.endswith(".xlsx"):
                file_path = os.path.join(self.output_directory, file)
                file_stats = os.stat(file_path)

                report_files.append(
                    {
                        "filename": file,
                        "filepath": file_path,
                        "size_mb": file_stats.st_size / (1024 * 1024),
                        "created": datetime.fromtimestamp(file_stats.st_ctime),
                        "modified": datetime.fromtimestamp(file_stats.st_mtime),
                    }
                )

        # Sort by creation time
        report_files.sort(
            key=lambda x: x["created"].timestamp()
            if hasattr(x["created"], "timestamp") and x["created"] is not None
            else 0.0,
            reverse=True,
        )

        return {
            "total_reports": len(report_files),
            "output_directory": self.output_directory,
            "recent_reports": report_files[:5],  # Most recent 5
            "total_size_mb": sum(f["size_mb"] for f in report_files),
        }
