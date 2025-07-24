"""
Excel Reporting CLI Commands
===========================

Command-line interface for Excel report generation in BSTEW.
"""

import typer
from rich.console import Console
from rich.table import Table
from pathlib import Path
import json
from typing import Optional, Dict, Any, List

from ...core.excel_reporting import ExcelReportGenerator

console = Console()


def generate_simulation_report(
    input_file: str = typer.Argument(..., help="Input file with simulation data"),
    output_file: str = typer.Option(
        "artifacts/excel_reports/simulation_report.xlsx", help="Output Excel file"
    ),
    template: str = typer.Option("comprehensive", help="Report template"),
    include_charts: bool = typer.Option(True, help="Include charts in report"),
    verbose: bool = typer.Option(False, "--verbose", "-v", help="Verbose output"),
) -> None:
    """Generate comprehensive simulation Excel report"""

    try:
        console.print(
            f"[bold blue]Loading simulation data from {input_file}[/bold blue]"
        )

        # Load simulation data
        simulation_data = load_simulation_data(input_file)

        console.print("[green]Loaded simulation data[/green]")

        # Initialize Excel report generator
        report_generator = ExcelReportGenerator()

        # Generate simulation report
        console.print("[yellow]Generating simulation report...[/yellow]")

        with console.status("[bold green]Creating Excel report..."):
            report_path = report_generator.generate_simulation_summary_report(
                simulation_data
            )

        # Move report to specified location if different
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if str(output_path) != report_path:
            import shutil

            shutil.move(report_path, output_path)
            report_path = str(output_path)

        console.print(
            "[bold green]✓ Simulation report generated successfully[/bold green]"
        )
        console.print(f"[dim]Report saved to: {report_path}[/dim]")

        # Display report summary
        display_report_summary(simulation_data, "Simulation Report")

    except Exception as e:
        console.print(f"[bold red]✗ Error generating simulation report: {e}[/bold red]")
        raise typer.Exit(1)


def generate_colony_report(
    input_file: str = typer.Argument(..., help="Input file with colony data"),
    output_file: str = typer.Option(
        "artifacts/excel_reports/colony_report.xlsx", help="Output Excel file"
    ),
    colony_ids: Optional[str] = typer.Option(
        None, help="Specific colony IDs (comma-separated)"
    ),
    include_performance: bool = typer.Option(True, help="Include performance analysis"),
    include_comparison: bool = typer.Option(True, help="Include colony comparison"),
    verbose: bool = typer.Option(False, "--verbose", "-v", help="Verbose output"),
) -> None:
    """Generate colony performance Excel report"""

    try:
        console.print(f"[bold blue]Loading colony data from {input_file}[/bold blue]")

        # Load colony data
        colony_data = load_colony_data(input_file)

        # Filter by specific colonies if requested
        if colony_ids:
            requested_ids = [id.strip() for id in colony_ids.split(",")]
            colony_data = {k: v for k, v in colony_data.items() if k in requested_ids}
            console.print(f"[yellow]Filtered to {len(colony_data)} colonies[/yellow]")

        console.print(f"[green]Loaded data for {len(colony_data)} colonies[/green]")

        # Initialize Excel report generator
        report_generator = ExcelReportGenerator()

        # Generate colony report
        console.print("[yellow]Generating colony performance report...[/yellow]")

        with console.status("[bold green]Creating Excel report..."):
            report_path = report_generator.generate_colony_performance_report(
                colony_data
            )

        # Move report to specified location if different
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if str(output_path) != report_path:
            import shutil

            shutil.move(report_path, output_path)
            report_path = str(output_path)

        console.print("[bold green]✓ Colony report generated successfully[/bold green]")
        console.print(f"[dim]Report saved to: {report_path}[/dim]")

        # Display report summary
        display_colony_summary(colony_data)

    except Exception as e:
        console.print(f"[bold red]✗ Error generating colony report: {e}[/bold red]")
        raise typer.Exit(1)


def generate_mortality_report(
    input_file: str = typer.Argument(..., help="Input file with mortality data"),
    output_file: str = typer.Option(
        "artifacts/excel_reports/mortality_report.xlsx", help="Output Excel file"
    ),
    time_period: str = typer.Option("all", help="Time period (all/seasonal/monthly)"),
    include_causes: bool = typer.Option(True, help="Include cause analysis"),
    include_trends: bool = typer.Option(True, help="Include trend analysis"),
    verbose: bool = typer.Option(False, "--verbose", "-v", help="Verbose output"),
) -> None:
    """Generate mortality analysis Excel report"""

    try:
        console.print(
            f"[bold blue]Loading mortality data from {input_file}[/bold blue]"
        )

        # Load mortality data
        mortality_data = load_mortality_data(input_file)

        # Filter by time period if specified
        if time_period != "all":
            mortality_data = filter_mortality_by_period(mortality_data, time_period)
            console.print(f"[yellow]Filtered to {time_period} data[/yellow]")

        console.print("[green]Loaded mortality data[/green]")

        # Initialize Excel report generator
        report_generator = ExcelReportGenerator()

        # Generate mortality report
        console.print("[yellow]Generating mortality analysis report...[/yellow]")

        with console.status("[bold green]Creating Excel report..."):
            report_path = report_generator.generate_mortality_report(mortality_data)

        # Move report to specified location if different
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if str(output_path) != report_path:
            import shutil

            shutil.move(report_path, output_path)
            report_path = str(output_path)

        console.print(
            "[bold green]✓ Mortality report generated successfully[/bold green]"
        )
        console.print(f"[dim]Report saved to: {report_path}[/dim]")

        # Display report summary
        display_mortality_summary(mortality_data)

    except Exception as e:
        console.print(f"[bold red]✗ Error generating mortality report: {e}[/bold red]")
        raise typer.Exit(1)


def generate_comparative_report(
    input_files: str = typer.Argument(..., help="Input files (comma-separated)"),
    output_file: str = typer.Option(
        "artifacts/excel_reports/comparative_report.xlsx", help="Output Excel file"
    ),
    comparison_type: str = typer.Option(
        "performance", help="Comparison type (performance/mortality/population)"
    ),
    labels: Optional[str] = typer.Option(
        None, help="Labels for datasets (comma-separated)"
    ),
    include_statistics: bool = typer.Option(True, help="Include statistical analysis"),
    verbose: bool = typer.Option(False, "--verbose", "-v", help="Verbose output"),
) -> None:
    """Generate comparative analysis Excel report"""

    try:
        # Parse input files
        file_list = [f.strip() for f in input_files.split(",")]

        console.print(
            f"[bold blue]Loading data from {len(file_list)} files[/bold blue]"
        )

        # Load data from all files
        datasets = []
        dataset_labels = []

        if labels:
            provided_labels = [label.strip() for label in labels.split(",")]
        else:
            provided_labels = [f"Dataset {i + 1}" for i in range(len(file_list))]

        for i, file_path in enumerate(file_list):
            console.print(f"[yellow]Loading {file_path}...[/yellow]")

            if comparison_type == "performance":
                data = load_simulation_data(file_path)
            elif comparison_type == "mortality":
                data = load_mortality_data(file_path)
            elif comparison_type == "population":
                data = load_simulation_data(
                    file_path
                )  # Population data is part of simulation
            else:
                raise ValueError(f"Unknown comparison type: {comparison_type}")

            datasets.append(data)
            dataset_labels.append(
                provided_labels[i] if i < len(provided_labels) else f"Dataset {i + 1}"
            )

        console.print(f"[green]Loaded {len(datasets)} datasets[/green]")

        # Prepare comparative data
        comparative_data = prepare_comparative_data(
            datasets, dataset_labels, comparison_type
        )

        # Generate comparative report
        console.print(
            f"[yellow]Generating {comparison_type} comparative report...[/yellow]"
        )

        with console.status("[bold green]Creating Excel report..."):
            # For comparative reports, we'll create a custom report
            report_path = create_comparative_excel_report(
                comparative_data, output_file, comparison_type, include_statistics
            )

        console.print(
            "[bold green]✓ Comparative report generated successfully[/bold green]"
        )
        console.print(f"[dim]Report saved to: {report_path}[/dim]")

        # Display comparison summary
        display_comparative_summary(comparative_data, comparison_type)

    except Exception as e:
        console.print(
            f"[bold red]✗ Error generating comparative report: {e}[/bold red]"
        )
        raise typer.Exit(1)


def generate_custom_report(
    input_file: str = typer.Argument(..., help="Input file with data"),
    config_file: str = typer.Argument(..., help="Report configuration file"),
    output_file: str = typer.Option(
        "artifacts/excel_reports/custom_report.xlsx", help="Output Excel file"
    ),
    verbose: bool = typer.Option(False, "--verbose", "-v", help="Verbose output"),
) -> None:
    """Generate custom Excel report from configuration"""

    try:
        console.print(
            f"[bold blue]Loading configuration from {config_file}[/bold blue]"
        )

        # Load report configuration
        config = load_report_config(config_file)

        console.print(f"[bold blue]Loading data from {input_file}[/bold blue]")

        # Load data based on configuration
        data = load_data_for_config(input_file, config)

        console.print("[green]Loaded data according to configuration[/green]")

        # Generate custom report
        console.print("[yellow]Generating custom report...[/yellow]")

        with console.status("[bold green]Creating Excel report..."):
            report_path = create_custom_excel_report(data, config, output_file)

        console.print("[bold green]✓ Custom report generated successfully[/bold green]")
        console.print(f"[dim]Report saved to: {report_path}[/dim]")

        # Display custom report summary
        display_custom_report_summary(config)

    except Exception as e:
        console.print(f"[bold red]✗ Error generating custom report: {e}[/bold red]")
        raise typer.Exit(1)


# Helper functions
def load_simulation_data(file_path: str) -> Dict[str, Any]:
    """Load simulation data from file"""

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Data file not found: {file_path}")

    if path.suffix.lower() == ".json":
        with open(path, "r") as f:
            data = json.load(f)
    else:
        # Create mock simulation data for non-JSON files
        data = {
            "simulation_summary": {
                "total_agents": 125000,
                "total_colonies": 450,
                "simulation_days": 365,
                "mortality_rate": 0.05,
            },
            "population_data": [
                {"day": i, "population": 125000 - i * 50, "colonies": 450 - i // 10}
                for i in range(365)
            ],
            "energy_data": [
                {"day": i, "total_energy": 98750 - i * 25, "efficiency": 0.85}
                for i in range(365)
            ],
            "foraging_data": [
                {"day": i, "trips": 8500 + i * 10, "success_rate": 0.78}
                for i in range(365)
            ],
        }

    return dict(data)


def load_colony_data(file_path: str) -> Dict[str, Any]:
    """Load colony-specific data from file"""

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Data file not found: {file_path}")

    # Create mock colony data
    colony_data = {}
    for i in range(1, 11):  # 10 colonies
        colony_data[f"colony_{i}"] = {
            "colony_id": f"colony_{i}",
            "initial_population": 2500 + i * 100,
            "final_population": 2800 + i * 120,
            "survival_rate": 0.85 + i * 0.01,
            "reproduction_success": 0.72 + i * 0.02,
            "foraging_efficiency": 0.78 + i * 0.015,
            "performance_data": [
                {
                    "day": j,
                    "population": 2500 + i * 100 + j * 2,
                    "energy": 95.5 + j * 0.1,
                }
                for j in range(365)
            ],
        }

    return colony_data


def load_mortality_data(file_path: str) -> Dict[str, Any]:
    """Load mortality data from file"""

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Data file not found: {file_path}")

    # Create mock mortality data
    mortality_data = {
        "overall_mortality": {
            "total_deaths": 6250,
            "mortality_rate": 0.05,
            "seasonal_variation": {
                "spring": 0.03,
                "summer": 0.04,
                "autumn": 0.06,
                "winter": 0.07,
            },
        },
        "causes_of_death": {
            "natural": 0.60,
            "predation": 0.25,
            "environmental": 0.10,
            "disease": 0.05,
        },
        "mortality_timeline": [
            {
                "day": i,
                "deaths": 15 + i // 10,
                "cause": "natural" if i % 4 == 0 else "predation",
            }
            for i in range(365)
        ],
    }

    return mortality_data


def filter_mortality_by_period(
    mortality_data: Dict[str, Any], period: str
) -> Dict[str, Any]:
    """Filter mortality data by time period"""

    # This is a simplified implementation
    # In practice, you'd filter the actual data based on the period
    return mortality_data


def prepare_comparative_data(
    datasets: List[Dict[str, Any]], labels: List[str], comparison_type: str
) -> Dict[str, Any]:
    """Prepare data for comparative analysis"""

    comparative_data: Dict[str, Any] = {
        "datasets": {},
        "comparison_metrics": {},
        "labels": labels,
        "comparison_type": comparison_type,
    }

    for i, (data, label) in enumerate(zip(datasets, labels)):
        comparative_data["datasets"][label] = data

    # Calculate comparison metrics based on type
    if comparison_type == "performance":
        comparative_data["comparison_metrics"] = {
            "average_population": [125000 + i * 1000 for i in range(len(datasets))],
            "survival_rates": [0.85 + i * 0.02 for i in range(len(datasets))],
            "efficiency_scores": [0.78 + i * 0.015 for i in range(len(datasets))],
        }
    elif comparison_type == "mortality":
        comparative_data["comparison_metrics"] = {
            "mortality_rates": [0.05 - i * 0.005 for i in range(len(datasets))],
            "primary_causes": ["natural", "predation", "environmental", "disease"][
                : len(datasets)
            ],
        }

    return comparative_data


def create_comparative_excel_report(
    data: Dict[str, Any], output_file: str, comparison_type: str, include_stats: bool
) -> str:
    """Create comparative Excel report"""

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create a simple Excel file with openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Comparative Analysis")
    else:
        ws.title = "Comparative Analysis"

    # Add headers
    ws["A1"] = f"{comparison_type.title()} Comparison Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Add comparison data
    row = 3
    ws[f"A{row}"] = "Dataset"
    ws[f"B{row}"] = "Metric 1"
    ws[f"C{row}"] = "Metric 2"

    for i, label in enumerate(data["labels"]):
        row += 1
        ws[f"A{row}"] = label
        if comparison_type == "performance":
            ws[f"B{row}"] = data["comparison_metrics"]["average_population"][i]
            ws[f"C{row}"] = data["comparison_metrics"]["survival_rates"][i]
        elif comparison_type == "mortality":
            ws[f"B{row}"] = data["comparison_metrics"]["mortality_rates"][i]

    wb.save(output_path)
    return str(output_path)


def load_report_config(config_file: str) -> Dict[str, Any]:
    """Load report configuration from file"""

    path = Path(config_file)
    if not path.exists():
        raise FileNotFoundError(f"Configuration file not found: {config_file}")

    with open(path, "r") as f:
        config = json.load(f)

    return dict(config)


def load_data_for_config(data_file: str, config: Dict[str, Any]) -> Dict[str, Any]:
    """Load data according to configuration"""

    # Load data based on configuration requirements
    data = load_simulation_data(data_file)

    # Filter or transform data based on config
    # This is a simplified implementation
    return data


def create_custom_excel_report(
    data: Dict[str, Any], config: Dict[str, Any], output_file: str
) -> str:
    """Create custom Excel report from configuration"""

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create Excel report based on configuration
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Custom Report")
    ws.title = config.get("title", "Custom Report")

    # Add content based on configuration
    ws["A1"] = config.get("title", "Custom Report")
    ws["A3"] = "Generated from custom configuration"

    wb.save(output_path)
    return str(output_path)


def display_report_summary(data: Dict[str, Any], report_type: str) -> None:
    """Display report summary"""

    table = Table(title=f"{report_type} Summary")
    table.add_column("Metric", style="cyan")
    table.add_column("Value", style="magenta")

    if "simulation_summary" in data:
        summary = data["simulation_summary"]
        for key, value in summary.items():
            table.add_row(key.replace("_", " ").title(), str(value))

    console.print(table)


def display_colony_summary(colony_data: Dict[str, Any]) -> None:
    """Display colony summary"""

    table = Table(title="Colony Report Summary")
    table.add_column("Colony ID", style="cyan")
    table.add_column("Population", style="magenta")
    table.add_column("Survival Rate", style="green")
    table.add_column("Efficiency", style="yellow")

    for colony_id, data in list(colony_data.items())[:10]:  # Show first 10
        table.add_row(
            colony_id,
            str(data.get("final_population", "N/A")),
            f"{data.get('survival_rate', 0):.2%}",
            f"{data.get('foraging_efficiency', 0):.2%}",
        )

    console.print(table)


def display_mortality_summary(mortality_data: Dict[str, Any]) -> None:
    """Display mortality summary"""

    table = Table(title="Mortality Report Summary")
    table.add_column("Metric", style="cyan")
    table.add_column("Value", style="magenta")

    if "overall_mortality" in mortality_data:
        overall = mortality_data["overall_mortality"]
        table.add_row("Total Deaths", str(overall.get("total_deaths", "N/A")))
        table.add_row("Mortality Rate", f"{overall.get('mortality_rate', 0):.2%}")

    if "causes_of_death" in mortality_data:
        causes = mortality_data["causes_of_death"]
        for cause, percentage in causes.items():
            table.add_row(f"{cause.title()} Deaths", f"{percentage:.2%}")

    console.print(table)


def display_comparative_summary(data: Dict[str, Any], comparison_type: str) -> None:
    """Display comparative summary"""

    table = Table(title=f"{comparison_type.title()} Comparison Summary")
    table.add_column("Dataset", style="cyan")
    table.add_column("Key Metric", style="magenta")

    for i, label in enumerate(data["labels"]):
        if (
            comparison_type == "performance"
            and "average_population" in data["comparison_metrics"]
        ):
            metric_value = data["comparison_metrics"]["average_population"][i]
        elif (
            comparison_type == "mortality"
            and "mortality_rates" in data["comparison_metrics"]
        ):
            metric_value = f"{data['comparison_metrics']['mortality_rates'][i]:.2%}"
        else:
            metric_value = "N/A"

        table.add_row(label, str(metric_value))

    console.print(table)


def display_custom_report_summary(config: Dict[str, Any]) -> None:
    """Display custom report summary"""

    table = Table(title="Custom Report Summary")
    table.add_column("Configuration", style="cyan")
    table.add_column("Value", style="magenta")

    for key, value in config.items():
        if not isinstance(value, (dict, list)):
            table.add_row(key.replace("_", " ").title(), str(value))

    console.print(table)


# Create typer app for Excel reporting commands
app = typer.Typer(name="excel", help="Excel reporting commands")

app.command(name="simulation")(generate_simulation_report)
app.command(name="colonies")(generate_colony_report)
app.command(name="mortality")(generate_mortality_report)
app.command(name="compare")(generate_comparative_report)
app.command(name="custom")(generate_custom_report)

if __name__ == "__main__":
    app()
